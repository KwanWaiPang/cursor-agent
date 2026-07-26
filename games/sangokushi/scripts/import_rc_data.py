#!/usr/bin/env python3
"""从 R-C-Group/shuju 与 touxiang 重建九州涂色武将数据。"""

from __future__ import annotations

import json
import re
import shutil
from pathlib import Path

import zhconv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "rc"
PORTRAIT_DST = ROOT / "assets" / "portraits"
SHUJU = Path("/tmp/shuju/san11")
TOUXIANG = Path("/tmp/touxiang/san/311_s")

VARIANT = str.maketrans(
    {
        "锺": "钟",
        "儁": "俊",
        "伷": "胄",
        "冑": "胄",
        "離": "离",
        "毗": "毘",
        "豫": "予",
        "麹": "麴",
        "関": "关",
        "霊": "灵",
        "観": "观",
        "楽": "乐",
        "徳": "德",
        "勲": "勋",
        "恵": "惠",
        "撃": "击",
        "戯": "戏",
        "単": "单",
        "厳": "严",
        "呉": "吴",
        "呂": "吕",
        "従": "从",
        "権": "权",
        "粛": "肃",
        "経": "经",
        "続": "续",
        "勵": "励",
        "塡": "填",
        "奬": "奖",
        "逹": "达",
        "邉": "边",
        "郷": "乡",
        "鎭": "镇",
        "髙": "高",
    }
)


def parse_chunks(path: Path):
    raw = path.read_text(encoding="utf-8")
    chunks = []
    depth = 0
    start = None
    for i, ch in enumerate(raw):
        if ch == "{":
            if depth == 0:
                start = i
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0 and start is not None:
                chunks.append(json.loads(raw[start : i + 1].replace("\r", "\\n")))
                start = None
    return chunks


def parse_array_file(path: Path):
    return json.loads(path.read_text(encoding="utf-8-sig").strip())


def norm_name(s: str) -> str:
    if s is None:
        return ""
    s = zhconv.convert(str(s), "zh-cn")
    s = s.replace(" ", "").replace("　", "")
    s = s.translate(VARIANT)
    # 去掉括号备注 张南(蜀)
    s = re.sub(r"[（(].*?[）)]", "", s)
    return s


def write_js(name: str, export_name: str, data):
    path = OUT / name
    path.write_text(
        "/** Auto-generated from R-C-Group — do not edit by hand */\n"
        f"export const {export_name} = {json.dumps(data, ensure_ascii=False, separators=(',', ':'))};\n",
        encoding="utf-8",
    )


def load_people_xlsx():
    wb = load_workbook(SHUJU / "人物.xlsx", read_only=True, data_only=True)
    rows = list(wb["原始数据"].iter_rows(values_only=True))
    wb.close()
    by_norm = {}
    for r in rows[1:]:
        vals = list(r) + [None] * 19
        name = vals[1]
        if not name:
            continue
        n = norm_name(name)
        by_norm[n] = {
            "affinity": vals[0],
            "trickName": vals[7],
            "birth": vals[14],
            "death": vals[15],
            "appear": vals[16],
            "home": norm_name(vals[17]) if vals[17] else "",
            "father": norm_name(vals[18]) if vals[18] else None,
            "rawName": str(name),
        }
    return by_norm


def load_relations_xlsx():
    wb = load_workbook(SHUJU / "关系.xlsx", read_only=True, data_only=True)
    rows = list(wb["Sheet1 (2)"].iter_rows(values_only=True))
    wb.close()
    officers = []
    cur = None
    phase = "like"
    for r in rows[4:]:
        vals = list(r) + [None] * 4
        a, n, m, t = vals[:4]
        if n:
            cur = {
                "name": norm_name(n),
                "affinity": a,
                "like": [],
                "hate": [],
                "bonds": [],
            }
            officers.append(cur)
            phase = "like"
        if cur is None:
            continue
        if t:
            tn = norm_name(t)
            cur[phase].append(tn)
            cur["bonds"].append(tn)
        elif a is None and n is None and m is None and t is None:
            if phase == "like" and cur["like"]:
                phase = "hate"
    return {o["name"]: o for o in officers}


def build_portrait_index():
    """文件名 -> Path，并建立规范化索引 / 异体映射。"""
    files = {p.stem: p for p in TOUXIANG.glob("*.jpg")}

    def find_pua(prefix: str):
        return [
            s
            for s in files
            if s.startswith(prefix) and any(0xE000 <= ord(c) <= 0xF8FF for c in s)
        ]

    def find_ext(prefix: str):
        return [s for s in files if s.startswith(prefix) and any(ord(c) > 0xFFFF for c in s)]

    explicit = {
        "车冑": "车胄",
        "荀勖": "荀勗",
        "田豫": "田予",
        "麴义": "麹义",
        "朱儁": "朱俊",
        "钟毓": "锺毓",
        "钟会": "锺会",
        "钟繇": "锺繇",
        "钟離牧": "锺离牧",
        "辛毗": "辛毘",
    }
    # 冷僻扩展区 / 错误编码为 PUA 的文件名
    if find_ext("荀"):
        # 荀𫖮
        for s in find_ext("荀"):
            if "𫖮" in s or any(ord(c) == 0x2B5AE for c in s):
                explicit["荀顗"] = s
    if find_ext("马云"):
        explicit["马云禄"] = find_ext("马云")[0]
    if find_pua("孔"):
        explicit["孔伷"] = find_pua("孔")[0]
    if find_pua("司马"):
        explicit["司马伷"] = find_pua("司马")[0]
    if find_pua("刘"):
        explicit["刘璝"] = find_pua("刘")[0]

    for k, v in list(explicit.items()):
        if v not in files:
            print("WARN explicit missing", k, repr(v))
            del explicit[k]
        else:
            print("map", k, "->", repr(v))
    return files, explicit


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    PORTRAIT_DST.mkdir(parents=True, exist_ok=True)
    for p in PORTRAIT_DST.glob("*.jpg"):
        p.unlink()

    gens = parse_chunks(SHUJU / "general.json")
    powers = parse_chunks(SHUJU / "power.json")
    tricks = parse_chunks(SHUJU / "trick.json")
    treasures = parse_chunks(SHUJU / "treasure.json")
    trick_types = parse_array_file(SHUJU / "trick_type.json")
    treasure_types = parse_array_file(SHUJU / "treasure_type.json")
    people = load_people_xlsx()
    relations = load_relations_xlsx()
    portrait_files, explicit = build_portrait_index()

    # norm portrait index
    norm_port = {}
    for stem in portrait_files:
        norm_port.setdefault(norm_name(stem), []).append(stem)

    name_to_id = {}
    officers = []
    missing_portraits = []
    portrait_map = {}

    for g in gens:
        oid = int(g["id"])
        name = g["name"]
        name_to_id[name] = oid
        name_to_id[norm_name(name)] = oid
        bio = (g.get("biography") or "").replace("\r\n", "\n").replace("\r", "\n")
        bio = re.sub(r"\n{3,}", "\n\n", bio).strip()
        person = people.get(norm_name(name), {})
        rel = relations.get(norm_name(name), {})

        # portrait
        stem = None
        if name in portrait_files:
            stem = name
        elif name in explicit and explicit[name] in portrait_files:
            stem = explicit[name]
        elif norm_name(name) in norm_port:
            stem = norm_port[norm_name(name)][0]
        else:
            missing_portraits.append(name)

        if stem:
            fname = f"{oid}.jpg"
            shutil.copy2(portrait_files[stem], PORTRAIT_DST / fname)
            portrait_map[str(oid)] = fname

        officers.append(
            {
                "id": oid,
                "name": name,
                "lead": int(g.get("command") or 0),
                "force": int(g.get("mforce") or 0),
                "int": int(g.get("intelligence") or 0),
                "pol": int(g.get("politics") or 0),
                "charm": int(g.get("charm") or 0),
                "trickId": int(g.get("trickId") if g.get("trickId") is not None else -1),
                "powerId": int(g.get("powerId") if g.get("powerId") is not None else 38),
                "apt": {
                    "gun": g.get("gun") or "C",
                    "halberd": g.get("halberd") or "C",
                    "crossbow": g.get("crossbow") or "C",
                    "ride": g.get("ride") or "C",
                    "weapons": g.get("weapons") or "C",
                    "water": g.get("water") or "C",
                },
                "bio": bio,
                "pic": g.get("pic") or "",
                "affinity": person.get("affinity"),
                "birth": person.get("birth"),
                "death": person.get("death"),
                "appear": person.get("appear"),
                "home": person.get("home") or "",
                "father": person.get("father"),
                "like": rel.get("like") or [],
                "hate": rel.get("hate") or [],
                "bonds": rel.get("bonds") or [],
            }
        )

    # resolve relation names -> ids
    for o in officers:
        def resolve(names):
            ids = []
            for n in names:
                nid = name_to_id.get(n) or name_to_id.get(norm_name(n))
                if nid is not None:
                    ids.append(nid)
            return ids

        o["likeIds"] = resolve(o["like"])
        o["hateIds"] = resolve(o["hate"])
        o["bondIds"] = resolve(o["bonds"])
        if o["father"]:
            o["fatherId"] = name_to_id.get(o["father"]) or name_to_id.get(norm_name(o["father"]))
        else:
            o["fatherId"] = None

    factions = []
    for p in powers:
        g = p.get("general") or {}
        factions.append(
            {
                "powerId": int(p["powerId"]),
                "name": g.get("name") or f"势力{p['powerId']}",
                "rulerId": int(g["id"]) if g.get("id") is not None else None,
            }
        )

    tricks_out = [
        {
            "id": int(t["id"]),
            "name": t["name"],
            "text": (t.get("text") or "").replace("\r\n", "\n").replace("\r", "\n"),
            "typeId": int(t.get("trickTypeId") or 0),
            "level": int(t.get("level") or 1),
            "arm": t.get("arm") or "",
        }
        for t in tricks
    ]
    treasures_out = []
    for t in treasures:
        gid = t.get("generalId")
        treasures_out.append(
            {
                "id": int(t["id"]),
                "name": t["name"],
                "typeId": int(t.get("typeId") or 0),
                "generalId": int(gid) if gid not in (None, "", 0, "0") else None,
                "desc": (t.get("state") or "").replace("\r\n", "\n").replace("\r", "\n").strip(),
                "pic": t.get("pic") or "",
            }
        )

    LEGACY = {
        "caocao": "曹操",
        "xiahoudun": "夏侯惇",
        "xiahouyuan": "夏侯渊",
        "caoren": "曹仁",
        "caoHong": "曹洪",
        "xuchu": "许褚",
        "dianwei": "典韦",
        "zhangliao": "张辽",
        "xuhuang": "徐晃",
        "zhanghe": "张郃",
        "yuejin": "乐进",
        "yujin": "于禁",
        "xunyu": "荀彧",
        "xunyou": "荀攸",
        "guojia": "郭嘉",
        "jiaxu": "贾诩",
        "simayi": "司马懿",
        "caopi": "曹丕",
        "caozhi": "曹植",
        "liubei": "刘备",
        "guanyu": "关羽",
        "zhangfei": "张飞",
        "zhaoyun": "赵云",
        "machao": "马超",
        "huangzhong": "黄忠",
        "zhugeliang": "诸葛亮",
        "pangtong": "庞统",
        "fazheng": "法正",
        "weiyan": "魏延",
        "jiangwei": "姜维",
        "liushan": "刘禅",
        "xushu": "徐庶",
        "humao": "马谡",
        "sunjian": "孙坚",
        "sunce": "孙策",
        "sunquan": "孙权",
        "zhouyu": "周瑜",
        "lusu": "鲁肃",
        "lvmeng": "吕蒙",
        "lvxun": "陆逊",
        "ganning": "甘宁",
        "taishici": "太史慈",
        "huanggai": "黄盖",
        "chengpu": "程普",
        "hanzang": "韩当",
        "zhoutai": "周泰",
        "yuanshao": "袁绍",
        "yanliang": "颜良",
        "wenchou": "文丑",
        "guoyuan": "郭图",
        "jushou": "沮授",
        "yuanShu": "袁术",
        "dongzhuo": "董卓",
        "lvbu": "吕布",
        "huaxiong": "华雄",
        "diaochan": "貂蝉",
        "gongsunzan": "公孙瓒",
        "liubiao": "刘表",
        "mateng": "马腾",
        "hanSui": "韩遂",
        "shixie": "士燮",
    }
    legacy_map = {slug: name_to_id[name] for slug, name in LEGACY.items() if name in name_to_id}

    # children index
    children = {}
    for o in officers:
        if o.get("fatherId"):
            children.setdefault(o["fatherId"], []).append(o["id"])
    for o in officers:
        o["childIds"] = children.get(o["id"], [])

    (OUT / "officers.json").write_text(
        json.dumps(officers, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    (OUT / "factions.json").write_text(json.dumps(factions, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "tricks.json").write_text(json.dumps(tricks_out, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "treasures.json").write_text(json.dumps(treasures_out, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "trick_types.json").write_text(json.dumps(trick_types, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "treasure_types.json").write_text(
        json.dumps(treasure_types, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT / "portraits.json").write_text(json.dumps(portrait_map, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT / "legacy_ids.json").write_text(json.dumps(legacy_map, ensure_ascii=False, indent=2), encoding="utf-8")

    write_js("officers_data.js", "RC_OFFICERS", officers)
    write_js("tricks_data.js", "RC_TRICKS", tricks_out)
    write_js("treasures_data.js", "RC_TREASURES", treasures_out)
    write_js("portraits_data.js", "RC_PORTRAITS", portrait_map)
    write_js("legacy_ids.js", "RC_LEGACY_IDS", legacy_map)
    write_js("power_factions.js", "RC_POWER_FACTIONS", factions)
    write_js("trick_types_data.js", "RC_TRICK_TYPES", trick_types)

    with_rel = sum(1 for o in officers if o["bondIds"] or o["fatherId"] or o["childIds"])
    with_father = sum(1 for o in officers if o["fatherId"])
    meta = {
        "source": "https://github.com/R-C-Group/shuju/tree/main/san11",
        "portraits": "https://github.com/R-C-Group/touxiang/tree/master/san/311_s",
        "officerCount": len(officers),
        "portraitCount": len(portrait_map),
        "missingPortraits": missing_portraits,
        "relationOfficers": len(relations),
        "officersWithRelations": with_rel,
        "officersWithFather": with_father,
        "powerCount": len(factions),
        "trickCount": len(tricks_out),
        "treasureCount": len(treasures_out),
    }
    (OUT / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(meta, ensure_ascii=False, indent=2))
    if missing_portraits:
        raise SystemExit(f"FAIL missing portraits: {missing_portraits}")
    print("OK all portraits matched")


if __name__ == "__main__":
    main()
